import requests, os, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_KEY = os.environ.get("API_KEY", "79c0592f98164ab5b336be10ef447644")
HEADERS = {"X-Auth-Token": API_KEY}
BASE    = "https://api.football-data.org/v4"

def thin():
    s = Side(border_style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_bottom():
    return Border(bottom=Side(border_style="medium", color="FFB0B0B0"))

# ── Data ophalen ───────────────────────────────────────────────────────────
print("Stand ophalen...")
stand_resp = requests.get(f"{BASE}/competitions/DED/standings", headers=HEADERS)
stand_resp.raise_for_status()
stand = stand_resp.json()["standings"][0]["table"]

print("Wedstrijden ophalen...")
match_resp = requests.get(f"{BASE}/competitions/DED/matches?status=FINISHED", headers=HEADERS)
match_resp.raise_for_status()
wedstrijden = match_resp.json()["matches"]

print(f"  → {len(stand)} teams, {len(wedstrijden)} wedstrijden")

# ── Teams op volgorde ──────────────────────────────────────────────────────
teams_stand  = [t["team"]["name"] for t in stand]          # op stand
teams_matrix = sorted(teams_stand)                          # alfabetisch voor matrix

# Resultaten-lookup: thuis[uit] = (thuis_goals, uit_goals)
resultaten = {}
for m in wedstrijden:
    th = m["homeTeam"]["name"]
    ut = m["awayTeam"]["name"]
    hg = m["score"]["fullTime"]["home"]
    ag = m["score"]["fullTime"]["away"]
    if hg is not None and ag is not None:
        resultaten.setdefault(th, {})[ut] = (hg, ag)

# ── Kleuren ────────────────────────────────────────────────────────────────
ORANJE   = "FFE87722"
DONKER   = "FF16213E"
WIT      = "FFFFFFFF"
LICHTGR  = "FFF5F5F5"
GOUD     = "FFFFD700"
ZILVER   = "FFD0D0D0"
BRONS    = "FFCD7F32"
BLAUW    = "FFE3F2FD"
ROOD_L   = "FFFDE8E8"
GROEN_L  = "FFE8F5E9"
GEEL_L   = "FFFFF9C4"
GRIJS_H  = "FFF0F0F0"

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# TABBLAD 1: KLASSEMENT
# ══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📊 Klassement"
ws.sheet_view.showGridLines = False

bijgewerkt = datetime.now().strftime("%-d %B %Y om %H:%M")

# Titel
ws.merge_cells("A1:K1")
ws["A1"] = f"⚽  EREDIVISIE 2025-2026  —  Bijgewerkt: {bijgewerkt}"
ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=WIT)
ws["A1"].fill      = PatternFill("solid", fgColor=DONKER)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Kolomkoppen
koppen = ["#", "Club", "G", "W", "GL", "V", "Pnt", "Voor", "Tgn", "Doel±", "W%"]
breedtes = [5, 24, 6, 6, 6, 6, 7, 7, 7, 7, 7]
for ci, (k, b) in enumerate(zip(koppen, breedtes), 1):
    c = ws.cell(row=2, column=ci, value=k)
    c.font      = Font(name="Arial", bold=True, size=10, color=WIT)
    c.fill      = PatternFill("solid", fgColor=ORANJE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()
    ws.column_dimensions[get_column_letter(ci)].width = b
ws.row_dimensions[2].height = 22

# Data rijen
for i, t in enumerate(stand):
    row = i + 3
    pos = t["position"]
    naam = t["team"]["name"]
    g    = t["playedGames"]
    w    = t["won"]
    gl   = t["draw"]
    v    = t["lost"]
    pnt  = t["points"]
    voor = t["goalsFor"]
    tgn  = t["goalsAgainst"]
    ds   = t["goalDifference"]
    wp   = round(w / g * 100) if g > 0 else 0

    # Achtergrond per zone
    if pos == 1:
        bg = "FFFFF8E1"   # goud-tint
    elif pos <= 3:
        bg = "FFF3F8FF"   # blauw-tint (Europa)
    elif pos <= 6:
        bg = LICHTGR
    elif pos >= 17:
        bg = "FFFFF0F0"   # rood-tint (degradatie)
    else:
        bg = WIT if i % 2 == 0 else GRIJS_H

    waarden = [pos, naam, g, w, gl, v, pnt, voor, tgn, ds, f"{wp}%"]
    for ci, val in enumerate(waarden, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.border    = thin()
        c.alignment = Alignment(
            horizontal="left" if ci == 2 else "center",
            vertical="center"
        )
        # Opmaak per kolom
        if ci == 1:   # positie
            c.font = Font(name="Arial", bold=True, size=10,
                          color="FF996600" if pos==1 else
                                "FF666666" if pos<=3 else "FF999999")
        elif ci == 2: # naam
            c.font = Font(name="Arial", bold=True, size=10)
        elif ci == 7: # punten
            c.font = Font(name="Arial", bold=True, size=11, color=ORANJE.replace("FF",""))
            c.font = Font(name="Arial", bold=True, size=11, color="E87722")
        elif ci == 10: # doelsaldo
            kleur = "2E7D32" if ds > 0 else "C62828" if ds < 0 else "555555"
            c.font = Font(name="Arial", bold=True, size=10, color=kleur)
            if ds > 0:
                c.value = f"+{ds}"
        else:
            c.font = Font(name="Arial", size=10)

    ws.row_dimensions[row].height = 20

# Legenda
leg_row = len(stand) + 4
ws.merge_cells(f"A{leg_row}:K{leg_row}")
ws[f"A{leg_row}"] = "🟡 Koploper   🔵 Europa (top 6)   🔴 Degradatie (positie 17-18)"
ws[f"A{leg_row}"].font      = Font(name="Arial", size=9, italic=True, color="FF888888")
ws[f"A{leg_row}"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[leg_row].height = 18

ws.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════
# TABBLAD 2: RESULTATEN MATRIX
# ══════════════════════════════════════════════════════════════════════════
wm = wb.create_sheet("⚽ Resultaten Matrix")
wm.sheet_view.showGridLines = False

n = len(teams_matrix)

# Titel
wm.merge_cells(f"A1:{get_column_letter(n+1)}1")
wm["A1"] = f"⚽  EREDIVISIE 2025-2026  —  Onderlinge Resultaten  —  Bijgewerkt: {bijgewerkt}"
wm["A1"].font      = Font(name="Arial", bold=True, size=12, color=WIT)
wm["A1"].fill      = PatternFill("solid", fgColor=DONKER)
wm["A1"].alignment = Alignment(horizontal="center", vertical="center")
wm.row_dimensions[1].height = 32

# Hoek + kolomkoppen (UIT-teams)
hoek = wm.cell(row=2, column=1, value="THUIS ↓ / UIT →")
hoek.font      = Font(name="Arial", bold=True, size=8, color=WIT)
hoek.fill      = PatternFill("solid", fgColor=DONKER)
hoek.alignment = Alignment(horizontal="center", vertical="bottom")
hoek.border    = thin()
wm.column_dimensions["A"].width = 20
wm.row_dimensions[2].height = 80

for ci, team in enumerate(teams_matrix, 2):
    c = wm.cell(row=2, column=ci, value=team)
    c.font      = Font(name="Arial", bold=True, size=8, color=WIT)
    c.fill      = PatternFill("solid", fgColor=ORANJE)
    c.alignment = Alignment(
        horizontal="center", vertical="bottom",
        text_rotation=90, wrap_text=True
    )
    c.border    = thin()
    wm.column_dimensions[get_column_letter(ci)].width = 7

# Rijen (THUIS-teams)
for ri, thuis in enumerate(teams_matrix):
    row = ri + 3

    # Rij-header
    rh = wm.cell(row=row, column=1, value=thuis)
    rh.font      = Font(name="Arial", bold=True, size=9)
    rh.fill      = PatternFill("solid", fgColor=ORANJE if ri % 2 == 0 else "FFCC6600")
    rh.fill      = PatternFill("solid", fgColor="FFE8E8E8" if ri % 2 == 0 else WIT)
    rh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    rh.border    = thin()
    wm.row_dimensions[row].height = 18

    for ci, uit in enumerate(teams_matrix, 2):
        c = wm.cell(row=row, column=ci)
        c.border    = thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font      = Font(name="Arial", bold=True, size=9)

        if thuis == uit:
            c.fill  = PatternFill("solid", fgColor="FF333333")
            c.value = "—"
            c.font  = Font(name="Arial", size=9, color="FF555555")
        elif thuis in resultaten and uit in resultaten[thuis]:
            hg, ag = resultaten[thuis][uit]
            c.value = f"{hg}-{ag}"
            if hg > ag:   # thuiswinst
                c.fill = PatternFill("solid", fgColor="FFD4EDDA")
                c.font = Font(name="Arial", bold=True, size=9, color="FF155724")
            elif hg < ag: # thuisverlies
                c.fill = PatternFill("solid", fgColor="FFF8D7DA")
                c.font = Font(name="Arial", bold=True, size=9, color="FF721C24")
            else:         # gelijk
                c.fill = PatternFill("solid", fgColor="FFFFF3CD")
                c.font = Font(name="Arial", bold=True, size=9, color="FF856404")
        else:
            c.value = "·"
            c.fill  = PatternFill("solid", fgColor="FFF8F8F8")
            c.font  = Font(name="Arial", size=9, color="FFCCCCCC")

# Legenda matrix
leg_m = len(teams_matrix) + 4
wm.merge_cells(f"A{leg_m}:{get_column_letter(n+1)}{leg_m}")
wm[f"A{leg_m}"] = "🟢 Groen = Thuiswinst   🔴 Rood = Thuisverlies   🟡 Geel = Gelijk   · = Nog niet gespeeld"
wm[f"A{leg_m}"].font      = Font(name="Arial", size=9, italic=True, color="FF888888")
wm[f"A{leg_m}"].alignment = Alignment(horizontal="left", vertical="center")
wm.row_dimensions[leg_m].height = 18

wm.freeze_panes = "B3"

# ══════════════════════════════════════════════════════════════════════════
# OPSLAAN
# ══════════════════════════════════════════════════════════════════════════
wb.save("Eredivisie_Live.xlsx")
print(f"✅  Eredivisie_Live.xlsx opgeslagen ({len(wedstrijden)} wedstrijden, {len(stand)} teams)")
